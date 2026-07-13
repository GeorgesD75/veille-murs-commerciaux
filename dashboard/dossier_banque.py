"""Dossiers de financement Excel « à présenter au banquier », un par annonce.

Un classeur openpyxl à 3 feuilles, pré-rempli avec les données de l'annonce :
- Synthèse : bandeau de marque, « l'essentiel » en 5 chiffres que le conseiller
  lit en 20 secondes, puis le bien, le marché local (ventes réelles DVF quand
  disponibles), les hypothèses (cellules jaunes MODIFIABLES), le plan de
  financement, l'exploitation et les ratios bancaires (DSCR, LTV) — le tout en
  FORMULES Excel : changer l'apport ou le taux recalcule tout, séance tenante ;
- Amortissement : tableau annuel indicatif + courbe du capital restant dû ;
- Annonce : les données brutes et sourcées (description, évolution du prix,
  alertes, score).

Honnêteté avant tout : un loyer estimé est marqué comme tel, un rendement
suspect garde son avertissement — un dossier gonflé se retourne contre vous.

Mise en page : les lignes sont posées par un CURSEUR et les formules relient
les cellules par leur position ENREGISTRÉE au moment de l'écriture (dict
`R`) — ajouter une ligne ne casse plus silencieusement les calculs.
"""
from __future__ import annotations

import re
import unicodedata
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from pipeline.config import Config
from pipeline.modeles import Annonce, TypeMurs

# --- style : sobre, lisible, imprimable ------------------------------------

_MARQUE = "1D5240"
_OR_VIF = "D6A532"
_TITRE_BLANC = Font(name="Calibri", size=17, bold=True, color="FFFFFF")
_SECTION = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_SECTION_FOND = PatternFill("solid", fgColor=_MARQUE)
_BANDEAU_FOND = PatternFill("solid", fgColor=_MARQUE)
_OR_FOND = PatternFill("solid", fgColor=_OR_VIF)
_LIBELLE = Font(name="Calibri", size=11, color="404040")
_VALEUR = Font(name="Calibri", size=11, bold=True)
_KPI_LIBELLE = Font(name="Calibri", size=11, bold=True, color=_MARQUE)
_KPI_VALEUR = Font(name="Calibri", size=14, bold=True)
_HYPOTHESE_FOND = PatternFill("solid", fgColor="FFF2CC")   # jaune : à ajuster
_NOTE = Font(name="Calibri", size=9, italic=True, color="808080")
_FILET = Border(bottom=Side(style="hair", color="D9D9D9"))
_VERT_FOND = PatternFill("solid", fgColor="E2EFDA")
_ROUGE_FOND = PatternFill("solid", fgColor="FCE4EC")

_EUR = '#,##0" €"'
_EUR_MOIS = '#,##0" €/mois"'
_PCT = "0.0%"
_M2 = '#,##0" m²"'

_LIGNES_AMORTISSEMENT = 30  # la durée est modifiable : IF() vide les années en trop


def _slug(texte: str) -> str:
    sans_accents = unicodedata.normalize("NFKD", texte).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]+", "-", sans_accents.lower()).strip("-") or "bien"


def _fmt_date(iso: str) -> str:
    return f"{iso[8:10]}/{iso[5:7]}" if len(iso) >= 10 else iso


def _evolution_prix_texte(a: Annonce) -> str | None:
    """« 420 000 € (03/07) → 380 000 € (12/07) — −9,5 % » ou None si stable."""
    hist = a.historique_prix or []
    if len(hist) < 2:
        return None
    etapes = " → ".join(
        f"{p['prix']:,.0f} € ({_fmt_date(p['date'])})".replace(",", " ") for p in hist
    )
    variation = (hist[-1]["prix"] / hist[0]["prix"] - 1) * 100
    signe = "−" if variation < 0 else "+"
    return f"{etapes}  —  {signe}{abs(variation):.1f} % depuis la première vue"


class _Feuille:
    """Curseur d'écriture : les positions clés sont ENREGISTRÉES, jamais devinées."""

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws
        self.r = 1
        self.R: dict[str, int] = {}   # nom logique -> numéro de ligne

    def sauter(self, n: int = 1) -> None:
        self.r += n

    def section(self, titre: str) -> None:
        ws = self.ws
        ws.merge_cells(start_row=self.r, start_column=1, end_row=self.r, end_column=3)
        c = ws.cell(self.r, 1, titre)
        c.font = _SECTION
        c.fill = _SECTION_FOND
        c.alignment = Alignment(indent=1, vertical="center")
        ws.row_dimensions[self.r].height = 18
        self.r += 1

    def ligne(self, cle: str | None, libelle: str, valeur: Any,
              fmt: str | None = None, hypothese: bool = False,
              note: str | None = None, police: Font | None = None) -> int:
        ws = self.ws
        ws.cell(self.r, 1, libelle).font = _LIBELLE
        cellule = ws.cell(self.r, 2, valeur)
        cellule.font = police or _VALEUR
        if fmt:
            cellule.number_format = fmt
        if hypothese:
            cellule.fill = _HYPOTHESE_FOND
        ws.cell(self.r, 1).border = _FILET
        cellule.border = _FILET
        if note:
            n = ws.cell(self.r, 3, note)
            n.font = _NOTE
            n.alignment = Alignment(vertical="center", wrap_text=True)
        if cle:
            self.R[cle] = self.r
        ligne = self.r
        self.r += 1
        return ligne

    def b(self, cle: str) -> str:
        """Référence Excel de la valeur enregistrée (colonne B)."""
        return f"B{self.R[cle]}"


def _params(config: Config) -> dict[str, Any]:
    analyse = config["analyse"]
    d = dict(analyse.get("dossier_banque") or {})
    financement = analyse.get("financement") or {}
    return {
        "seuil_score": d.get("seuil_score", 60),
        "apport_pct": d.get("apport_pct", 20),
        "frais_acquisition_pct": d.get("frais_acquisition_pct", 8),
        "taxe_fonciere_mois_loyer": d.get("taxe_fonciere_mois_loyer", 1),
        "provision_entretien_pct": d.get("provision_entretien_pct", 5),
        "assurance_pno_eur_an": d.get("assurance_pno_eur_an", 300),
        "taux_pct": financement.get("taux_pct", 3.9),
        "duree_ans": financement.get("duree_ans", 20),
    }


def _verdict(annonce: Annonce, p: dict[str, Any]) -> str:
    """Phrase honnête pour la 1re page, aux hypothèses PAR DÉFAUT du classeur."""
    loyer = annonce.loyer_mensuel or annonce.loyer_mensuel_estime
    if not loyer or not annonce.prix:
        return ("Loyer inconnu à ce stade : le dossier est pré-rempli côté "
                "acquisition ; renseignez le loyer (feuille Synthèse) pour "
                "activer tous les calculs.")
    cout = annonce.prix * (1 + p["frais_acquisition_pct"] / 100)
    emprunt = cout * (1 - p["apport_pct"] / 100)
    t, n = p["taux_pct"] / 100 / 12, p["duree_ans"] * 12
    mensualite = emprunt * t / (1 - (1 + t) ** -n)
    charges = (loyer * p["taxe_fonciere_mois_loyer"] / 12
               + loyer * p["provision_entretien_pct"] / 100
               + p["assurance_pno_eur_an"] / 12)
    cf = loyer - mensualite - charges
    cf_txt = f"{abs(cf):,.0f}".replace(",", " ")
    taux_txt = str(p["taux_pct"]).replace(".", ",")
    prefixe = ("Sur la base d'un loyer ESTIMÉ (pas de bail signé à ce stade) : "
               if annonce.loyer_estime or not annonce.loyer_mensuel else "")
    hypotheses = (f"apport {p['apport_pct']} %, crédit {taux_txt} % "
                  f"sur {p['duree_ans']} ans, charges provisionnées")
    if cf >= 0:
        return (f"{prefixe}aux hypothèses de ce dossier ({hypotheses}), le bien "
                f"s'autofinance : le loyer couvre crédit et charges et laisse "
                f"environ +{cf_txt} €/mois.")
    return (f"{prefixe}aux hypothèses de ce dossier ({hypotheses}), il reste "
            f"environ {cf_txt} €/mois à financer en plus du loyer — "
            f"marge de négociation ou apport supplémentaire à prévoir.")


# --- feuille 1 : Synthèse ----------------------------------------------------

def _feuille_synthese(ws: Worksheet, a: Annonce, p: dict[str, Any]) -> dict[str, int]:
    ws.title = "Synthèse"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 17
    ws.column_dimensions["C"].width = 46

    # Bandeau de marque + filet doré : le classeur se présente tout seul.
    ws.merge_cells("A1:C1")
    ws["A1"] = "Les Murs.   Dossier de financement — murs commerciaux"
    ws["A1"].font = _TITRE_BLANC
    ws["A1"].fill = _BANDEAU_FOND
    ws["A1"].alignment = Alignment(vertical="center", indent=1)
    ws.row_dimensions[1].height = 34
    ws.merge_cells("A2:C2")
    ws["A2"].fill = _OR_FOND
    ws.row_dimensions[2].height = 3

    ws.merge_cells("A3:C3")
    ws["A3"] = a.titre
    ws["A3"].font = Font(size=13, bold=True)
    ws.merge_cells("A4:B4")
    ws["A4"] = f"{a.ville} ({a.code_postal}) — annonce {a.source} du {a.date_premiere_vue[:10]}"
    ws["A4"].font = _LIBELLE
    lien = ws.cell(4, 3, "Voir l'annonce en ligne ↗")
    lien.hyperlink = a.url
    lien.font = Font(color="0563C1", underline="single", size=10)
    ws.cell(5, 1, f"Dossier généré le {date.today():%d/%m/%Y} — indicatif, l'offre de prêt de la banque fait foi.").font = _NOTE

    f = _Feuille(ws)
    f.r = 7

    # « L'ESSENTIEL » est rempli à la FIN (ses formules pointent vers des
    # cellules écrites plus bas) : on réserve les lignes maintenant.
    ligne_essentiel = f.r
    f.sauter(8)

    f.section("LE BIEN")
    etat = ("Murs occupés — un locataire en place paie déjà un loyer"
            if a.type_murs is TypeMurs.MURS_OCCUPES
            else "Murs libres — local à louer après l'achat")
    f.ligne(None, "État locatif", etat)
    f.ligne("surface", "Surface", a.surface_m2, _M2)
    f.ligne("prix_affiche", "Prix affiché", a.prix, _EUR)
    f.ligne(None, "Prix au m²", f"=IFERROR({f.b('prix_affiche')}/{f.b('surface')},\"\")", _EUR)
    note_loyer = None
    if a.loyer_mensuel:
        note_loyer = "loyer du bail en place (à faire confirmer par le bail)"
    elif a.loyer_mensuel_estime:
        note_loyer = "ESTIMATION au loyer de marché local — aucun bail ne le garantit"
    f.ligne("loyer", "Loyer mensuel HT", a.loyer_mensuel or a.loyer_mensuel_estime,
            _EUR, hypothese=a.loyer_mensuel is None, note=note_loyer)
    if a.marche_prix_m2_bas and a.marche_prix_m2_haut:
        f.ligne(
            None, "Marché local (prix/m²)",
            f"{a.marche_prix_m2_bas:,.0f} à {a.marche_prix_m2_haut:,.0f} €/m²".replace(",", " "),
            note=(f"ce bien est {abs(a.decote_pct):.0f} % "
                  f"{'SOUS' if a.decote_pct >= 0 else 'au-dessus de'} la médiane locale"
                  if a.decote_pct is not None else None),
        )
        if a.benchmark_source:
            # Une fourchette assise sur des ventes NOTARIÉES (DVF) pèse plus
            # lourd dans une négociation bancaire qu'un référentiel maison.
            f.ligne(None, "Source de la fourchette", a.benchmark_source,
                    note="DVF = prix réellement payés, enregistrés chez le notaire (data.gouv.fr)"
                    if "DVF" in a.benchmark_source else None)
    evolution = _evolution_prix_texte(a)
    if evolution:
        f.ligne(None, "Évolution du prix demandé", evolution,
                note="un prix qui baisse est un levier de négociation documenté")
    if a.dpe_classe:
        f.ligne(None, "Classe énergie (DPE)", a.dpe_classe,
                note="annoncée par le vendeur — demander le diagnostic complet")
    f.sauter()

    f.section("HYPOTHÈSES — cellules jaunes à ajuster avec votre conseiller")
    f.ligne("prix_negocie", "Prix négocié retenu", a.prix, _EUR, hypothese=True,
            note="par défaut le prix affiché — un prix demandé se négocie")
    f.ligne("frais_pct", "Frais d'acquisition (notaire, garanties)",
            p["frais_acquisition_pct"] / 100, _PCT, hypothese=True)
    f.ligne("apport_pct", "Apport personnel", p["apport_pct"] / 100, _PCT, hypothese=True)
    f.ligne("taux", "Taux du crédit (hors assurance)", p["taux_pct"] / 100, _PCT, hypothese=True)
    f.ligne("duree", "Durée du crédit (années)", p["duree_ans"], "0", hypothese=True)
    f.ligne("tf_mois", "Taxe foncière (en mois de loyer par an)",
            p["taxe_fonciere_mois_loyer"], "0.0", hypothese=True,
            note="à remplacer par le montant réel dès que connu")
    f.ligne("entretien_pct", "Provision entretien & vacance (% du loyer)",
            p["provision_entretien_pct"] / 100, _PCT, hypothese=True)
    f.ligne("pno", "Assurance propriétaire (PNO, €/an)",
            p["assurance_pno_eur_an"], _EUR, hypothese=True)
    f.sauter()

    f.section("PLAN DE FINANCEMENT")
    f.ligne("cout_total", "Coût d'acquisition total",
            f"={f.b('prix_negocie')}*(1+{f.b('frais_pct')})", _EUR,
            note="prix négocié + frais d'acquisition")
    f.ligne("apport", "Apport personnel", f"={f.b('cout_total')}*{f.b('apport_pct')}", _EUR)
    f.ligne("emprunt", "Montant emprunté", f"={f.b('cout_total')}-{f.b('apport')}", _EUR)
    f.ligne("mensualite", "Mensualité du crédit",
            f"=IFERROR(-PMT({f.b('taux')}/12,{f.b('duree')}*12,{f.b('emprunt')}),0)", _EUR_MOIS)
    f.ligne("service_dette", "Service de la dette (annuel)", f"={f.b('mensualite')}*12", _EUR)
    f.sauter()

    f.section("EXPLOITATION ANNUELLE")
    f.ligne("loyers_an", "Loyers annuels HT", f"=IFERROR({f.b('loyer')}*12,0)", _EUR)
    f.ligne("tf", "− Taxe foncière", f"=-IFERROR({f.b('loyer')}*{f.b('tf_mois')},0)", _EUR)
    f.ligne("entretien", "− Entretien & vacance locative",
            f"=-{f.b('loyers_an')}*{f.b('entretien_pct')}", _EUR)
    f.ligne("assurance", "− Assurance PNO", f"=-{f.b('pno')}", _EUR)
    f.ligne("revenu_net", "Revenu net avant crédit",
            f"={f.b('loyers_an')}+{f.b('tf')}+{f.b('entretien')}+{f.b('assurance')}", _EUR)
    f.ligne(None, "− Remboursement du crédit", f"=-{f.b('service_dette')}", _EUR)
    f.ligne("cf_an", "Cash-flow annuel", f"={f.b('revenu_net')}-{f.b('service_dette')}", _EUR)
    f.ligne("cf_mois", "Cash-flow mensuel", f"={f.b('cf_an')}/12", _EUR_MOIS,
            note="positif : le bien se paie tout seul, loyer ≥ crédit + charges",
            police=Font(size=13, bold=True))
    for cle in ("cf_an", "cf_mois"):
        ws.conditional_formatting.add(f.b(cle), CellIsRule(
            operator="greaterThanOrEqual", formula=["0"], fill=_VERT_FOND))
        ws.conditional_formatting.add(f.b(cle), CellIsRule(
            operator="lessThan", formula=["0"], fill=_ROUGE_FOND))
    f.sauter()

    f.section("LES RATIOS QUE REGARDE LE BANQUIER")
    f.ligne(None, "Rendement brut", f"=IFERROR({f.b('loyers_an')}/{f.b('prix_negocie')},\"\")",
            _PCT, note="loyer annuel ÷ prix — la base de comparaison des biens")
    f.ligne(None, "Rendement net avant crédit",
            f"=IFERROR({f.b('revenu_net')}/{f.b('cout_total')},\"\")", _PCT,
            note="après taxe foncière, entretien et assurance")
    f.ligne("dscr", "DSCR — couverture de la dette",
            f"=IFERROR({f.b('revenu_net')}/{f.b('service_dette')},\"\")", "0.00",
            note="revenu net ÷ crédit : ≥ 1,20 rassure une banque")
    ws.conditional_formatting.add(f.b("dscr"), CellIsRule(
        operator="greaterThanOrEqual", formula=["1.2"], fill=_VERT_FOND))
    ws.conditional_formatting.add(f.b("dscr"), CellIsRule(
        operator="lessThan", formula=["1"], fill=_ROUGE_FOND))
    f.ligne(None, "LTV — part financée par la banque",
            f"=IFERROR({f.b('emprunt')}/{f.b('prix_negocie')},\"\")", _PCT,
            note="≤ 80 % est la zone de confort bancaire")
    f.ligne(None, "Effort d'épargne mensuel si négatif", f"=MAX(0,-{f.b('cf_an')}/12)", _EUR_MOIS)
    f.ligne(None, "Score de l'outil de veille",
            f"{a.score}/100" if a.score is not None else "—",
            note="rendement, emplacement, prix vs marché, financement, fiscalité")

    if "rendement_anormalement_eleve" in (a.flags or []):
        f.sauter()
        ws.merge_cells(start_row=f.r, start_column=1, end_row=f.r + 1, end_column=3)
        alerte = ws.cell(f.r, 1,
                         "⚠ AVERTISSEMENT : le rendement affiché par le vendeur est "
                         "anormalement élevé. Exiger le bail et les quittances avant "
                         "toute offre — ce dossier reprend le loyer annoncé tel quel.")
        alerte.alignment = Alignment(wrap_text=True, vertical="top")
        alerte.font = Font(size=10, bold=True, color="98351B")
        f.sauter(2)

    derniere = f.r

    # --- L'ESSENTIEL, rempli maintenant que toutes les positions sont connues.
    ws.merge_cells(start_row=ligne_essentiel, start_column=1,
                   end_row=ligne_essentiel, end_column=3)
    c = ws.cell(ligne_essentiel, 1, "L'ESSENTIEL — ce que votre conseiller lit en premier")
    c.font = _SECTION
    c.fill = _SECTION_FOND
    c.alignment = Alignment(indent=1, vertical="center")
    ws.row_dimensions[ligne_essentiel].height = 18
    kpis = [
        ("Prix d'achat retenu", f"={f.b('prix_negocie')}", _EUR),
        ("Apport nécessaire", f"={f.b('apport')}", _EUR),
        ("Mensualité de crédit estimée", f"={f.b('mensualite')}", _EUR_MOIS),
        ("Cash-flow mensuel attendu", f"={f.b('cf_mois')}", _EUR_MOIS),
        ("DSCR (≥ 1,20 rassure)", f"={f.b('dscr')}", "0.00"),
    ]
    for i, (libelle, formule, fmt) in enumerate(kpis):
        r = ligne_essentiel + 1 + i
        ws.cell(r, 1, libelle).font = _KPI_LIBELLE
        cellule = ws.cell(r, 2, formule)
        cellule.font = _KPI_VALEUR
        cellule.number_format = fmt
        ws.cell(r, 1).border = _FILET
        cellule.border = _FILET
    ws.conditional_formatting.add(f"B{ligne_essentiel + 4}", CellIsRule(
        operator="greaterThanOrEqual", formula=["0"], fill=_VERT_FOND))
    ws.conditional_formatting.add(f"B{ligne_essentiel + 4}", CellIsRule(
        operator="lessThan", formula=["0"], fill=_ROUGE_FOND))
    ws.merge_cells(start_row=ligne_essentiel + 6, start_column=1,
                   end_row=ligne_essentiel + 7, end_column=3)
    verdict = ws.cell(ligne_essentiel + 6, 1, "EN CLAIR — " + _verdict(a, p))
    verdict.alignment = Alignment(wrap_text=True, vertical="top")
    verdict.font = Font(size=10.5, italic=True)

    # petit bloc de données pour le graphique (à droite, hors zone d'impression)
    ws["E7"] = "Loyer mensuel"
    ws["F7"] = f"=IFERROR({f.b('loyer')},0)"
    ws["E8"] = "Crédit + charges / mois"
    ws["F8"] = f"={f.b('mensualite')}-({f.b('tf')}+{f.b('entretien')}+{f.b('assurance')})/12"
    ws["E9"] = "Cash-flow"
    ws["F9"] = f"={f.b('cf_mois')}"
    for r in (7, 8, 9):
        ws.cell(r, 5).font = _NOTE
        ws.cell(r, 6).number_format = _EUR

    graphique = BarChart()
    graphique.type = "col"
    graphique.title = "Le loyer paie-t-il le crédit ?"
    graphique.legend = None
    graphique.y_axis.numFmt = _EUR
    graphique.height, graphique.width = 8, 12
    graphique.add_data(Reference(ws, min_col=6, min_row=7, max_row=9))
    graphique.set_categories(Reference(ws, min_col=5, min_row=7, max_row=9))
    graphique.series[0].graphicalProperties.solidFill = _MARQUE
    ws.add_chart(graphique, "E11")

    # impression : A4 portrait, ajusté en largeur, pied de page sobre
    ws.print_area = f"A1:C{derniere}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.oddFooter.center.text = "Les Murs. — dossier indicatif généré automatiquement ; l'offre de prêt de la banque fait foi"
    ws.oddFooter.center.size = 8

    return f.R


# --- feuille 2 : Amortissement ----------------------------------------------

def _feuille_amortissement(ws: Worksheet, R: dict[str, int]) -> None:
    ws.title = "Amortissement"
    ws.sheet_view.showGridLines = False
    entetes = ["Année", "Capital dû en début d'année", "Intérêts de l'année",
               "Capital remboursé", "Capital restant dû"]
    for col, entete in enumerate(entetes, start=1):
        c = ws.cell(1, col, entete)
        c.font = _SECTION
        c.fill = _SECTION_FOND
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[chr(64 + col)].width = 22
    ws.cell(2, 6, "Tableau indicatif (pas annuel) — l'offre de prêt de la banque fait foi.").font = _NOTE

    ws["G1"] = f"='Synthèse'!B{R['duree']}"
    ws["G2"] = f"='Synthèse'!B{R['taux']}"
    ws["G3"] = f"='Synthèse'!B{R['service_dette']}"
    ws.column_dimensions["G"].hidden = True
    for i in range(_LIGNES_AMORTISSEMENT):
        r = i + 2
        ws.cell(r, 1, i + 1)
        debut = f"'Synthèse'!$B${R['emprunt']}" if i == 0 else f"E{r - 1}"
        garde = f"A{r}<=$G$1"
        ws.cell(r, 2, f"=IF({garde},{debut},\"\")")
        ws.cell(r, 3, f"=IF({garde},B{r}*$G$2,\"\")")
        ws.cell(r, 4, f"=IF({garde},MIN(B{r},$G$3-C{r}),\"\")")
        ws.cell(r, 5, f"=IF({garde},MAX(0,B{r}-D{r}),\"\")")
        for col in (2, 3, 4, 5):
            ws.cell(r, col).number_format = _EUR

    courbe = LineChart()
    courbe.title = "Capital restant dû"
    courbe.legend = None
    courbe.y_axis.numFmt = _EUR
    courbe.x_axis.title = "Année"
    courbe.height, courbe.width = 9, 16
    courbe.add_data(Reference(ws, min_col=5, min_row=2,
                              max_row=_LIGNES_AMORTISSEMENT + 1))
    courbe.set_categories(Reference(ws, min_col=1, min_row=2,
                                    max_row=_LIGNES_AMORTISSEMENT + 1))
    courbe.series[0].graphicalProperties.line.solidFill = _MARQUE
    courbe.series[0].graphicalProperties.line.width = 22000  # ~1,75 pt
    ws.add_chart(courbe, "I2")


# --- feuille 3 : Annonce (les faits, sourcés) --------------------------------

def _feuille_annonce(ws: Worksheet, a: Annonce) -> None:
    ws.title = "Annonce"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 100
    lignes: list[tuple[str, Any]] = [
        ("Titre", a.titre),
        ("Lien", a.url),
        ("Ville", f"{a.ville} ({a.code_postal})"),
        ("Source", a.source),
        ("Première détection", a.date_premiere_vue[:10]),
        ("Type", "Murs occupés" if a.type_murs is TypeMurs.MURS_OCCUPES else "Murs libres"),
        ("Caractéristiques", " · ".join(a.caracteristiques) or "—"),
        ("Évolution du prix", _evolution_prix_texte(a) or "stable depuis la première vue"),
        ("Fourchette de marché", a.benchmark_source or "—"),
        ("Lecture du prix", a.lecture_prix or "—"),
        ("Alertes", " · ".join(a.flags) or "aucune"),
        ("Description", a.description or "—"),
    ]
    for r, (libelle, valeur) in enumerate(lignes, start=1):
        ws.cell(r, 1, libelle).font = Font(bold=True)
        cellule = ws.cell(r, 2, valeur)
        cellule.alignment = Alignment(wrap_text=True, vertical="top")
        if libelle == "Lien":
            cellule.hyperlink = a.url
            cellule.font = Font(color="0563C1", underline="single")
    ws.row_dimensions[len(lignes)].height = 110


# --- point d'entrée -----------------------------------------------------------

def generer_dossier(annonce: Annonce, config: Config, cible: Path) -> None:
    """Écrit le classeur d'UNE annonce (utile aux tests)."""
    p = _params(config)
    wb = Workbook()
    positions = _feuille_synthese(wb.active, annonce, p)
    _feuille_amortissement(wb.create_sheet(), positions)
    _feuille_annonce(wb.create_sheet(), annonce)
    cible.parent.mkdir(parents=True, exist_ok=True)
    wb.save(cible)


def generer_dossiers(
    annonces: dict[str, Annonce], config: Config, dossier: Path
) -> dict[str, str]:
    """Génère les classeurs des annonces retenues à score ≥ seuil.

    Retourne {id annonce: nom de fichier} pour les liens du dashboard.
    Le répertoire est reconstruit à chaque run : pas de dossier fantôme
    pointant vers une annonce disparue.
    """
    seuil = _params(config)["seuil_score"]
    dossier.mkdir(parents=True, exist_ok=True)
    for ancien in dossier.glob("*.xlsx"):
        ancien.unlink()
    resultats: dict[str, str] = {}
    for a in annonces.values():
        if a.exclue or a.prix is None or (a.score or 0) < seuil:
            continue
        nom = f"dossier-{_slug(a.ville)}-{a.id[:8]}.xlsx"
        generer_dossier(a, config, dossier / nom)
        resultats[a.id] = nom
    return resultats
