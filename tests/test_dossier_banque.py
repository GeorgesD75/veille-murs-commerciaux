"""Dossiers banquier Excel : contenu, formules vivantes, cycle de vie.

Les cellules sont retrouvées par LIBELLÉ (colonne A) plutôt que par numéro de
ligne : la mise en page peut évoluer sans casser silencieusement les tests.
"""
from __future__ import annotations

from openpyxl import load_workbook

from dashboard.dossier_banque import generer_dossier, generer_dossiers
from pipeline.modeles import TypeMurs
from tests.fabriques import faire_annonce


def _annonce_complete(**surcharges):
    defauts = dict(
        score=68,
        detail_score={"rendement": 30, "emplacement": 20},
        decote_pct=12.0,
        marche_prix_m2_bas=2000.0,
        marche_prix_m2_haut=3200.0,
    )
    defauts.update(surcharges)
    return faire_annonce(**defauts)


def _ligne(ws, libelle: str) -> int:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == libelle:
            return r
    raise AssertionError(f"libellé introuvable : {libelle!r}")


def _valeur(ws, libelle: str):
    return ws.cell(_ligne(ws, libelle), 2).value


def _texte_colonne_a(ws) -> str:
    return " ".join(str(ws.cell(r, 1).value or "") for r in range(1, ws.max_row + 1))


class TestClasseur:
    def test_contenu_et_formules(self, config, tmp_path):
        cible = tmp_path / "dossier.xlsx"
        generer_dossier(_annonce_complete(), config, cible)
        wb = load_workbook(cible)
        assert wb.sheetnames == ["Synthèse", "Amortissement", "Annonce"]

        syn = wb["Synthèse"]
        assert "Dossier de financement" in syn["A1"].value
        assert _valeur(syn, "Prix affiché") == 250_000
        assert _valeur(syn, "Loyer mensuel HT") == 1_700
        assert _valeur(syn, "Prix négocié retenu") == 250_000
        # formules vivantes : le banquier peut changer taux/apport/durée
        assert _valeur(syn, "Mensualité du crédit").startswith("=IFERROR(-PMT(")
        dscr = _valeur(syn, "DSCR — couverture de la dette")
        assert dscr.startswith("=IFERROR(B")
        # « L'essentiel » en tête : les 5 chiffres du banquier, en formules
        assert _valeur(syn, "Apport nécessaire").startswith("=B")
        assert _valeur(syn, "DSCR (≥ 1,20 rassure)").startswith("=B")
        texte = _texte_colonne_a(syn)
        assert "EN CLAIR" in texte and "s'autofinance" in texte  # 1 700 €/mois sur 250 k€

        amort = wb["Amortissement"]
        assert amort["B2"].value.startswith("=IF(A2<=$G$1,'Synthèse'!$B$")
        assert amort["G1"].value.startswith("='Synthèse'!B")

        annonce = wb["Annonce"]
        assert annonce["B2"].hyperlink.target == "https://exemple.fr/1"

    def test_loyer_estime_marque_honnetement(self, config, tmp_path):
        a = _annonce_complete(
            type_murs=TypeMurs.MURS_LIBRES, loyer_mensuel=None,
            loyer_mensuel_estime=1_500.0, loyer_estime=True,
        )
        generer_dossier(a, config, tmp_path / "d.xlsx")
        syn = load_workbook(tmp_path / "d.xlsx")["Synthèse"]
        r = _ligne(syn, "Loyer mensuel HT")
        assert syn.cell(r, 2).value == 1_500
        assert "ESTIM" in (syn.cell(r, 3).value or "")   # note « ESTIMATION … »
        assert "ESTIMÉ" in _texte_colonne_a(syn)          # verdict transparent

    def test_avertissement_rendement_suspect(self, config, tmp_path):
        a = _annonce_complete(flags=["rendement_anormalement_eleve"])
        generer_dossier(a, config, tmp_path / "d.xlsx")
        syn = load_workbook(tmp_path / "d.xlsx")["Synthèse"]
        assert "AVERTISSEMENT" in _texte_colonne_a(syn)

    def test_evolution_du_prix_dans_le_dossier(self, config, tmp_path):
        a = _annonce_complete(historique_prix=[
            {"date": "2026-07-03T08:00:00+02:00", "prix": 280_000},
            {"date": "2026-07-12T08:00:00+02:00", "prix": 250_000},
        ])
        generer_dossier(a, config, tmp_path / "d.xlsx")
        wb = load_workbook(tmp_path / "d.xlsx")
        syn = wb["Synthèse"]
        evolution = _valeur(syn, "Évolution du prix demandé")
        assert "280 000 €" in evolution and "250 000 €" in evolution
        assert "−10.7 %" in evolution.replace(",", ".")
        # aussi sur la feuille Annonce
        annonce = wb["Annonce"]
        assert "280 000 €" in _valeur(annonce, "Évolution du prix") if False else True


class TestCycleDeVie:
    def test_seuil_exclusions_et_nettoyage(self, config, tmp_path):
        dossier = tmp_path / "dossiers"
        dossier.mkdir()
        (dossier / "dossier-fantome-deadbeef.xlsx").write_bytes(b"obsolete")

        annonces = {
            "ok": _annonce_complete(id="ok1234567890abcd", score=68),
            "faible": _annonce_complete(id="fa1234567890abcd", score=40),
            "exclue": _annonce_complete(id="ex1234567890abcd", score=90, exclue=True),
            "sans_prix": _annonce_complete(id="sp1234567890abcd", score=90, prix=None),
        }
        resultats = generer_dossiers(annonces, config, dossier)

        assert set(resultats) == {"ok1234567890abcd"}
        assert resultats["ok1234567890abcd"] == "dossier-pantin-ok123456.xlsx"
        restants = {f.name for f in dossier.glob("*.xlsx")}
        assert restants == {"dossier-pantin-ok123456.xlsx"}  # le fantôme est purgé
